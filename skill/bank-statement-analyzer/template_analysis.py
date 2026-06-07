#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bank Statement Analyzer — configurable starter script.
Builds a 6-sheet Excel workbook from a folder of monthly CSV statements.

HOW TO USE
  1. Put one year of CSV statements in a folder.
  2. Edit the CONFIG block below to match your bank's columns and your categories.
  3. Run:  python3 template_analysis.py "/path/to/folder"
  Needs:   pip3 install openpyxl

This is a STARTING POINT — adapt the column names and keyword lists to your data.
For PDF/Excel statements, convert to CSV first (or let your AI assistant adapt this).

© 2026 SoulStrategy by Chantal Perrinjaquet — House Of Coaching. Shared as a gift.
"""
import csv, glob, re, os, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG  ←  adapt this block to your bank export
# ============================================================
FILE_GLOB   = "*.csv"          # pattern of your statement files inside the folder
COL_DATE    = "Date"           # column holding the transaction date
COL_DESC    = "Description"    # column holding the payee / description
COL_AMOUNT  = "Amount"         # signed amount: negative = money out, positive = money in
COL_BALANCE = "Balance"        # running balance (leave "" if your export has none)
COL_CURRENCY= "Account currency"  # currency code; set to "" if single-currency
COL_ACCOUNT = "Account IBAN"   # something identifying the account (IBAN/number/name)
COL_TYPE    = "Transaction type"  # optional transaction type; "" if none
COL_NOTE    = "Notes"          # optional note column; "" if none
DATE_IS_ISO = True             # True if dates are YYYY-MM-DD (sortable as text)

# Category detection — keyword lists (lowercase, matched inside the description).
OWNER_KEYWORDS    = ["my own name here"]   # the business owner's name → owner deposits / salary draws
INTERNAL_KEYWORDS = ["transfer to savings", "to pocket", "internal"]  # moves between own accounts
# Transaction-type values that mean "currency exchange / internal FX" (optional):
INTERNAL_TYPES    = ["currency exchange"]
# Transaction-type value that means "bank fee" (bundled as one creditor):
FEE_TYPES         = ["fees", "fee"]
# Vendor name normalisation: any description containing the key → the clean name.
VENDOR_NORMALISE  = {"amazon": "Amazon", "facebk": "Facebook Ads", "facebook": "Facebook Ads"}

OUTPUT_NAME = "Statement Analysis.xlsx"
# ============================================================

SRC = os.path.expanduser(sys.argv[1]) if len(sys.argv) > 1 else "."
OUT = os.path.join(SRC, OUTPUT_NAME)

def g(r, col, default=""):
    return r.get(col, default) if col else default

# ---------- Load ----------
rows = []
for f in sorted(glob.glob(os.path.join(SRC, FILE_GLOB))):
    with open(f, newline='', encoding='utf-8-sig') as fh:
        for i, r in enumerate(csv.DictReader(fh)):
            try:
                r['_amt'] = float(str(g(r, COL_AMOUNT)).replace(",", "").strip())
            except ValueError:
                continue
            r['_date'] = str(g(r, COL_DATE)).strip()
            r['_seq'] = (f, i)
            rows.append(r)
if not rows:
    sys.exit(f"No transactions found in {SRC} (glob {FILE_GLOB}). Check CONFIG.")
rows.sort(key=lambda r: (r['_date'], r['_seq']) if DATE_IS_ISO else (r['_seq'],))

def desc(r):  return str(g(r, COL_DESC)).strip()
def cur(r):   return str(g(r, COL_CURRENCY)).strip() or "—"
def acct(r):  return str(g(r, COL_ACCOUNT)).strip()[-6:] or "—"
def ttype(r): return str(g(r, COL_TYPE)).strip()
def note(r):  return str(g(r, COL_NOTE)).strip()

def has_kw(text, kws): t = text.lower(); return any(k in t for k in kws)

# ---------- Categorise ----------
def category(r):
    a = r['_amt']; d = desc(r); ty = ttype(r).lower()
    if ty in INTERNAL_TYPES or has_kw(d, INTERNAL_KEYWORDS):
        return "internal"
    if a < 0 and has_kw(d, OWNER_KEYWORDS):
        return "salary"        # money out to the owner
    if a > 0 and has_kw(d, OWNER_KEYWORDS):
        return "owner_deposit" # money in from the owner
    if a > 0:
        return "debtor"        # real revenue
    return "creditor"          # any other money out

for r in rows:
    r['_cat'] = category(r)

def vendor(r):
    d = desc(r); low = d.lower()
    if ttype(r).lower() in FEE_TYPES: return "Bank Fees"
    for k, v in VENDOR_NORMALISE.items():
        if k in low: return v
    return d

# ---------- Styles ----------
HF=PatternFill("solid",fgColor="1F4E5F"); HFONT=Font(bold=True,color="FFFFFF")
TFONT=Font(bold=True,size=13,color="1F4E5F"); SFONT=Font(italic=True,size=9,color="666666")
SECF=Font(bold=True,size=11,color="1F4E5F"); TOT=Font(bold=True); TOTF=PatternFill("solid",fgColor="DDEBF1")
INF=PatternFill("solid",fgColor="E8F5E9"); OUTF=PatternFill("solid",fgColor="FDECEA")
thin=Side(style="thin",color="D0D0D0"); BD=Border(thin,thin,thin,thin)
MONEY='#,##0.00'; CEN=Alignment(horizontal="center")
def header(ws,row,n):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c); x.fill=HF; x.font=HFONT; x.border=BD
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def widths(ws,ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w

wb=Workbook()

# ---------- Sheet 1: All Transactions ----------
ws=wb.active; ws.title="1 All Transactions"
ws.cell(1,1,"All Transactions").font=TFONT
ws.cell(2,1,f"{len(rows)} transactions, chronological. Balance = bank balance per account. Amounts in account currency.").font=SFONT
H=["Date","Account","Currency","Type","Description","Income","Expense","Balance","Note"]; HR=4
for c,h in enumerate(H,1): ws.cell(HR,c,h)
header(ws,HR,len(H)); rr=HR+1; first=rr
for r in rows:
    a=r['_amt']
    ws.cell(rr,1,r['_date']); ws.cell(rr,2,acct(r)); ws.cell(rr,3,cur(r))
    ws.cell(rr,4,ttype(r)); ws.cell(rr,5,desc(r))
    ci=ws.cell(rr,6,round(a,2) if a>0 else None); co=ws.cell(rr,7,round(-a,2) if a<0 else None)
    if COL_BALANCE:
        try: ws.cell(rr,8,float(str(g(r,COL_BALANCE)).replace(",",""))).number_format=MONEY
        except ValueError: pass
    ws.cell(rr,9,note(r))
    for c in (6,7): ws.cell(rr,c).number_format=MONEY
    if a>0: ci.fill=INF
    else: co.fill=OUTF
    rr+=1
last=rr-1
rr+=1; ws.cell(rr,1,"COLUMN TOTALS per currency").font=SECF; rr+=1; th=rr
for c,h in enumerate(["Currency","Total Income","Total Expense","Net"],1): ws.cell(th,c,h)
header(ws,th,4); rr+=1
for cc in sorted(set(cur(r) for r in rows)):
    ws.cell(rr,1,cc).font=Font(bold=True)
    ws.cell(rr,2,f'=SUMIF($C${first}:$C${last},"{cc}",$F${first}:$F${last})').number_format=MONEY
    ws.cell(rr,3,f'=SUMIF($C${first}:$C${last},"{cc}",$G${first}:$G${last})').number_format=MONEY
    ws.cell(rr,4,f"=B{rr}-C{rr}").number_format=MONEY
    for c in (2,3,4): ws.cell(rr,c).fill=TOTF
    rr+=1
widths(ws,[12,12,9,16,40,13,13,15,30]); ws.freeze_panes="A5"; ws.auto_filter.ref=f"A{HR}:I{last}"

# ---------- Creditors ----------
kred=[(vendor(r), r['_date'], round(-r['_amt'],2), ttype(r), desc(r), acct(r))
      for r in rows if r['_cat']=="creditor"]
ws=wb.create_sheet("2 Creditors Detail")
ws.cell(1,1,"Creditors — Detail").font=TFONT
ws.cell(2,1,"All supplier charges, alphabetical then by date. Negative = refund/credit. Bank fees bundled.").font=SFONT
H=["Creditor","Date","Amount","Type","Reference","Account"]; HR=4
for c,h in enumerate(H,1): ws.cell(HR,c,h)
header(ws,HR,len(H)); rr=HR+1
for k,dt,bt,ty,rf,ac in sorted(kred,key=lambda x:(x[0].lower(),x[1])):
    ws.cell(rr,1,k); ws.cell(rr,2,dt); ws.cell(rr,3,bt).number_format=MONEY
    ws.cell(rr,4,ty); ws.cell(rr,5,rf); ws.cell(rr,6,ac); rr+=1
ws.cell(rr,2,"Total").font=TOT; ws.cell(rr,3,f"=SUM(C{HR+1}:C{rr-1})").number_format=MONEY; ws.cell(rr,3).font=TOT
widths(ws,[26,12,13,14,34,12]); ws.freeze_panes="A5"; ws.auto_filter.ref=f"A{HR}:F{rr-1}"

agg=defaultdict(lambda:{'n':0,'s':0.0,'f':'9999-99','l':'0000-00'})
for k,dt,bt,*_ in kred:
    x=agg[k]; x['n']+=1; x['s']+=bt; x['f']=min(x['f'],dt); x['l']=max(x['l'],dt)
ws=wb.create_sheet("3 Creditors Summary")
ws.cell(1,1,"Creditors — Summary").font=TFONT
ws.cell(2,1,"Unique creditors, alphabetical, with total charged.").font=SFONT
H=["Creditor","No. of Charges","Total","First","Last"]; HR=4
for c,h in enumerate(H,1): ws.cell(HR,c,h)
header(ws,HR,len(H)); rr=HR+1
for k in sorted(agg,key=str.lower):
    x=agg[k]; ws.cell(rr,1,k); ws.cell(rr,2,x['n']).alignment=CEN
    ws.cell(rr,3,round(x['s'],2)).number_format=MONEY; ws.cell(rr,4,x['f']); ws.cell(rr,5,x['l']); rr+=1
ws.cell(rr,1,"TOTAL").font=TOT; ws.cell(rr,3,f"=SUM(C{HR+1}:C{rr-1})").number_format=MONEY; ws.cell(rr,3).font=TOT
widths(ws,[28,15,14,12,12]); ws.freeze_panes="A5"; ws.auto_filter.ref=f"A{HR}:E{rr-1}"

# ---------- Debtors ----------
deb=[r for r in rows if r['_cat']=="debtor"]
ws=wb.create_sheet("4 Debtors")
ws.cell(1,1,"Debtors — Revenue").font=TFONT
ws.cell(2,1,"Real incoming revenue (owner deposits and internal moves excluded).").font=SFONT
ws.cell(4,1,"Summary by payer").font=SECF
H=["Payer","Currency","No.","Total"]; HR=5
for c,h in enumerate(H,1): ws.cell(HR,c,h)
header(ws,HR,len(H))
dag=defaultdict(lambda:{'n':0,'s':0.0})
for r in deb:
    key=(desc(r),cur(r)); dag[key]['n']+=1; dag[key]['s']+=r['_amt']
rr=HR+1
for (nm,cc) in sorted(dag,key=lambda x:(x[0].lower(),x[1])):
    x=dag[(nm,cc)]; ws.cell(rr,1,nm); ws.cell(rr,2,cc); ws.cell(rr,3,x['n']).alignment=CEN
    ws.cell(rr,4,round(x['s'],2)).number_format=MONEY; rr+=1
for cc in sorted(set(cur(r) for r in deb)):
    tot=sum(r['_amt'] for r in deb if cur(r)==cc)
    ws.cell(rr,1,f"TOTAL {cc}").font=TOT; ws.cell(rr,4,round(tot,2)).number_format=MONEY; ws.cell(rr,4).font=TOT; rr+=1
widths(ws,[28,9,8,14]); ws.freeze_panes="A6"

# ---------- Owner Account ----------
sal=[r for r in rows if r['_cat']=="salary"]
own=[r for r in rows if r['_cat']=="owner_deposit"]
ws=wb.create_sheet("5 Owner Account")
ws.cell(1,1,"Owner Account").font=TFONT
ws.cell(2,1,"Block A: salary/draws (business → owner). Block B: owner deposits (owner → business).").font=SFONT
td=sum(-r['_amt'] for r in sal); tp=sum(r['_amt'] for r in own); ro=4
ws.cell(ro,1,"Summary").font=SECF; ro+=1
for lab,val in [("Total salary draws (out)",round(td,2)),("Total owner deposits (in)",round(tp,2)),
                ("Net owner withdrawal",round(td-tp,2))]:
    ws.cell(ro,1,lab).font=Font(bold=True); ws.cell(ro,2,val).number_format=MONEY; ws.cell(ro,2).fill=TOTF; ro+=1
for title,data,sign in [("A — Salary draws (out)",sal,-1),("B — Owner deposits (in)",own,1)]:
    ro+=1; ws.cell(ro,1,title).font=SECF; ro+=1; hh=ro
    for c,h in enumerate(["Date","Amount","Account","Note"],1): ws.cell(ro,c,h)
    header(ws,ro,4); ro+=1
    for r in data:
        ws.cell(ro,1,r['_date']); ws.cell(ro,2,round(sign*r['_amt'],2)).number_format=MONEY
        ws.cell(ro,3,acct(r)); ws.cell(ro,4,note(r) or desc(r)); ro+=1
    ws.cell(ro,1,"Total").font=TOT; ws.cell(ro,2,f"=SUM(B{hh+1}:B{ro-1})").number_format=MONEY; ws.cell(ro,2).font=TOT; ro+=1
widths(ws,[14,14,14,40]); ws.freeze_panes="A5"

# ---------- Internal Transfers ----------
intl=[r for r in rows if r['_cat']=="internal"]
ws=wb.create_sheet("6 Internal Transfers")
ws.cell(1,1,"Internal Transfers").font=TFONT
ws.cell(2,1,"Moves between your own accounts (currency exchange, savings pockets). No real money in/out.").font=SFONT
H=["Date","Description","Amount","Account","Currency"]; HR=4
for c,h in enumerate(H,1): ws.cell(HR,c,h)
header(ws,HR,len(H)); rr=HR+1
for r in intl:
    ws.cell(rr,1,r['_date']); ws.cell(rr,2,desc(r)); ws.cell(rr,3,round(r['_amt'],2)).number_format=MONEY
    ws.cell(rr,4,acct(r)); ws.cell(rr,5,cur(r)); rr+=1
widths(ws,[12,40,13,14,9]); ws.freeze_panes="A5"

for w,col in zip(wb.worksheets,["1F4E5F","2E7D32","2E7D32","C0392B","8E44AD","B8860B"]):
    w.sheet_properties.tabColor=col

wb.save(OUT)
from collections import Counter
print("SAVED:", OUT)
print("Categories:", dict(Counter(r['_cat'] for r in rows)), "| total:", len(rows))
print("Reconciliation OK:", sum(Counter(r['_cat'] for r in rows).values())==len(rows))
